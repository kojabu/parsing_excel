import openpyxl
import requests
import os
import re

# Specify the correct filename with .xlsx extension
wb = openpyxl.load_workbook('Inventory.xlsx')

# Check available sheets
print("Available sheets:", wb.sheetnames)

# Open the first sheet
ws = wb.active
print(f"Working with sheet: {ws.title}")

# Regular expression to extract URL from formula
url_pattern = re.compile(r'"(http?://[^"]+)"')  # Match text inside quotes starting with http/https

# Folder for saving images
output_dir = 'output'
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Iterate through rows 2–202 and columns P (16) – BB (67)
for row in range(2, 203):  # Rows 2–202
    barcode_cell = ws.cell(row=row, column=6)  # Barcode in column F (6)
    barcode = str(barcode_cell.value).strip()  # Get the product barcode
    
    if not barcode:  # Skip if the barcode is empty
        continue

    # Create a folder for the product if it doesn't exist
    barcode_folder = os.path.join(output_dir, barcode)
    if not os.path.exists(barcode_folder):
        os.makedirs(barcode_folder)

    # Iterate through columns with hyperlinks to images (P-BB), from 16th to 68th
    image_counter = 1
    for col in range(16, 68):  # Columns P (16) – BB (67)
        cell = ws.cell(row=row, column=col)
        value = cell.value
        if value and isinstance(value, str) and value.startswith("=HYPERLINK"):  # Check for HYPERLINK formula
            match = url_pattern.search(value)  # Search for URL in the string
            if match:
                url = match.group(1)  # Extract URL from regex group
                print(f"Downloading: {url}")

                # File name for saving
                file_name = os.path.join(barcode_folder, f"{barcode}-{image_counter}.jpg")
                
                try:
                    # Download the image
                    response = requests.get(url, stream=True, verify=False)  
                    if response.status_code == 200:
                        # Save the image
                        with open(file_name, 'wb') as file:
                            file.write(response.content)
                        print(f"File saved: {file_name}")
                    else:
                        print(f"Error downloading {url}: {response.status_code}")
                except Exception as e:
                    print(f"Error downloading {url}: {e}")

                # Increment the image counter
                image_counter += 1
            else:
                print(f"Failed to extract link from: {value}")

print("Download complete.")
